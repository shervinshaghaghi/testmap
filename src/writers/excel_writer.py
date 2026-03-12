# src/writers/excel_writer.py
from pathlib import Path
import re
from collections import Counter
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.worksheet import Worksheet


_REQ_RE = re.compile(r"\bR(\d{1,2})\b", re.IGNORECASE)


def _reqs_from_requirement_id(rid: str, max_r: int = 19) -> List[str]:
    """
    Returns list of requirements counted for this row.
    - "R5" -> ["R5"]
    - "OOB" -> []
    - "MIX: R3+R7" -> ["R3","R7"]
    """
    rid = (rid or "").strip()
    up = rid.upper()

    if not up or up == "OOB":
        return []

    nums = [int(m.group(1)) for m in _REQ_RE.finditer(up)]
    out = []
    for n in nums:
        if 1 <= n <= max_r:
            out.append(f"R{n}")
    # de-dup but keep order
    seen = set()
    dedup = []
    for r in out:
        if r not in seen:
            seen.add(r)
            dedup.append(r)
    return dedup


def compute_counts(result: Dict[str, Any], max_r: int = 19) -> Dict[str, int]:
    rows = result.get("rows") or []
    counts = Counter()
    oob = 0

    for r in rows:
        rid = str(r.get("requirement_id", "") or "").strip()
        up = rid.upper()

        if not up or up == "OOB":
            oob += 1
            continue

        for req in _reqs_from_requirement_id(up, max_r=max_r):
            counts[req] += 1

    out = {f"R{i}": int(counts.get(f"R{i}", 0)) for i in range(1, max_r + 1)}
    out["OOB"] = int(oob)
    out["TOTAL_CASES"] = int(len(rows))
    return out


def _autosize(ws: Worksheet) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 55)


def _style_header(ws: Worksheet) -> None:
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def write_student_xlsx(result: Dict[str, Any], out_path: Path, max_r: int = 19) -> None:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # --- Sheet 1: Mapping ---
    ws = wb.active
    ws.title = "Mapping"
    ws.append(["Student", "CaseID", "RequirementID", "Rationale"])
    _style_header(ws)

    student = (result.get("student") or "").strip() or "STUDENT"
    for r in (result.get("rows") or []):
        ws.append([
            student,
            str(r.get("case_id", "") or ""),
            str(r.get("requirement_id", "") or ""),
            str(r.get("rationale", "") or ""),
        ])
    _autosize(ws)

    # --- Sheet 2: Counts ---
    ws2 = wb.create_sheet("Counts")
    header = ["CO/IN", "Student"] + [f"R{i}" for i in range(1, max_r + 1)] + ["Out of the box"]
    ws2.append(header)
    _style_header(ws2)

    c = compute_counts(result, max_r=max_r)
    row = ["", student] + [c[f"R{i}"] for i in range(1, max_r + 1)] + [c["OOB"]]
    ws2.append(row)

    # center numeric cells
    for col in range(3, 3 + max_r + 1):
        ws2.cell(row=2, column=col).alignment = Alignment(horizontal="center")

    _autosize(ws2)
    wb.save(out_path)


def write_batch_counts_xlsx(results: List[Dict[str, Any]], out_path: Path, max_r: int = 19) -> None:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Counts"

    header = ["CO/IN", "Student"] + [f"R{i}" for i in range(1, max_r + 1)] + ["Out of the box"]
    ws.append(header)
    _style_header(ws)

    for result in results:
        student = (result.get("student") or "").strip() or "STUDENT"
        c = compute_counts(result, max_r=max_r)
        ws.append(["", student] + [c[f"R{i}"] for i in range(1, max_r + 1)] + [c["OOB"]])

    # center numeric cells
    for r in range(2, ws.max_row + 1):
        for col in range(3, 3 + max_r + 1):
            ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")

    _autosize(ws)
    wb.save(out_path)